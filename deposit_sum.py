import calendar
import openpyxl as op
from openpyxl.styles import Font


def deposit_sum():
    load_excel = op.load_workbook('summary.xlsx')
    data_sheet = load_excel["by currency"]
    font = Font(u'Tahoma', size=8)
    if "deposit" in load_excel.sheetnames:
        sheet = load_excel["deposit"]
    else:
        sheet = load_excel.create_sheet("deposit")
        sheet.column_dimensions['A'].width = 20.0
        sheet.column_dimensions['B'].width = 20.0
        sheet.column_dimensions['C'].width = 20.0
        sheet.column_dimensions['D'].width = 20.0
        sheet.column_dimensions['H'].width = 12.0
        sheet.column_dimensions['I'].width = 12.0
        sheet.column_dimensions['J'].width = 12.0
        sheet.column_dimensions['K'].width = 12.0
        sheet['A1'] = 'DayOfMonth'
        sheet['B1'] = 'User'
        sheet['C1'] = 'Count'
        sheet['D1'] = 'Volume'
        sheet['H16'] = 'Total'
        sheet['H18'] = 'W1'
        sheet['H19'] = 'W2'
        sheet['H20'] = 'W3'
        sheet['H21'] = 'W4'
        sheet['H24'] = 'Avg'
        sheet['H25'] = 'W1'
        sheet['H26'] = 'W2'
        sheet['H27'] = 'W3'
        sheet['H28'] = 'W4'
        sheet['I15'] = 'User'
        sheet['I16'] = '=SUM(B2:B32)'
        sheet['I24'] = '=AVERAGE(B2:B32)'
        sheet['J15'] = 'Count'
        sheet['J16'] = '=SUM(C2:C32)'
        sheet['J24'] = '=AVERAGE(C2:C32)'
        sheet['K15'] = 'Volume'
        sheet['K16'] = '=SUM(D2:D32)'
        sheet['K24'] = '=AVERAGE(D2:D32)'

    start = 0
    user = 0
    count = 0
    amount = 0.00

    for i in range(2, data_sheet.max_row + 2):
        if start == 0:
            start = int(data_sheet["A" + str(i)].value)
        if data_sheet["A" + str(i)].value is None:
            sheet["A" + str(start + 1)] = start
            sheet["B" + str(start + 1)] = user
            sheet["C" + str(start + 1)] = count
            sheet["D" + str(start + 1)] = amount
            break
        elif int(data_sheet["A" + str(i)].value) == start:
            if data_sheet["B" + str(i)].value is not None:
                user += int(data_sheet["C" + str(i)].value)
            if data_sheet["C" + str(i)].value is not None:
                count += int(data_sheet["D" + str(i)].value)
            if data_sheet["D" + str(i)].value is not None:
                amount += float(data_sheet["E" + str(i)].value)
        elif int(data_sheet["A" + str(i)].value) > start:
            sheet["A" + str(start + 1)] = start
            sheet["B" + str(start + 1)] = user
            sheet["C" + str(start + 1)] = count
            sheet["D" + str(start + 1)] = amount
            start = int(data_sheet["A" + str(i)].value)
            user = int(data_sheet["C" + str(i)].value)
            count = int(data_sheet["D" + str(i)].value)
            amount = float(data_sheet["E" + str(i)].value)

    for row in range(2, sheet.max_row + 1):
        sheet["B" + str(row)].number_format = '#,##'
        sheet["C" + str(row)].number_format = '#,##'
        sheet["D" + str(row)].number_format = '#,##0.00'

    print("please insert year: ", end="")
    year = input()
    print("please insert month: ", end="")
    month = input()
    cal = calendar.weekday(int(year), int(month), 1)

    if cal >= 2:
        first_wednesday_date = cal
    else:
        first_wednesday_date = cal + 1

    current_count_date = first_wednesday_date + 1

    for w in range(0, 4):
        if first_wednesday_date == 1:
            sheet["I18"] = '=SUM(B' + str(current_count_date) + ':B' + str(current_count_date + 6) + ')'
            sheet["J18"] = '=SUM(C' + str(current_count_date) + ':C' + str(current_count_date + 6) + ')'
            sheet["K18"] = '=SUM(D' + str(current_count_date) + ':D' + str(current_count_date + 6) + ')'
            sheet["I25"] = '=AVERAGE(B' + str(current_count_date) + ':B' + str(current_count_date + 6) + ')'
            sheet["J25"] = '=AVERAGE(C' + str(current_count_date) + ':C' + str(current_count_date + 6) + ')'
            sheet["K25"] = '=AVERAGE(D' + str(current_count_date) + ':D' + str(current_count_date + 6) + ')'
        elif w == 0:
            sheet["I18"] = '-'
            sheet["J18"] = '-'
            sheet["K18"] = '-'
            sheet["I25"] = '-'
            sheet["J25"] = '-'
            sheet["K25"] = '-'
            continue
        sheet["I" + str(18 + w)] = '=SUM(B' + str(current_count_date) + ':B' + str(current_count_date + 6) + ')'
        sheet["J" + str(18 + w)] = '=SUM(C' + str(current_count_date) + ':C' + str(current_count_date + 6) + ')'
        sheet["K" + str(18 + w)] = '=SUM(D' + str(current_count_date) + ':D' + str(current_count_date + 6) + ')'
        sheet["I" + str(25 + w)] = '=AVERAGE(B' + str(current_count_date) + ':B' + str(current_count_date + 6) + ')'
        sheet["J" + str(25 + w)] = '=AVERAGE(C' + str(current_count_date) + ':C' + str(current_count_date + 6) + ')'
        sheet["K" + str(25 + w)] = '=AVERAGE(D' + str(current_count_date) + ':D' + str(current_count_date + 6) + ')'
        current_count_date += 7

    for math_format in range(16, 30):
        if math_format == 22 or math_format == 29:
            sheet["I" + str(math_format)].number_format = '0.00%'
            sheet["J" + str(math_format)].number_format = '0.00%'
            sheet["K" + str(math_format)].number_format = '0.00%'
            continue
        sheet["I" + str(math_format)].number_format = '#,##'
        sheet["J" + str(math_format)].number_format = '#,##'
        sheet["K" + str(math_format)].number_format = '#,##'

    for font_format in range(1, sheet.max_row + 1):
        sheet['A' + str(font_format)].font = font
        sheet['B' + str(font_format)].font = font
        sheet['C' + str(font_format)].font = font
        sheet['D' + str(font_format)].font = font
        sheet['H' + str(font_format)].font = font
        sheet['I' + str(font_format)].font = font
        sheet['J' + str(font_format)].font = font
        sheet['K' + str(font_format)].font = font

    load_excel.save('summary.xlsx')
