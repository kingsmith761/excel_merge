import os
import openpyxl
from openpyxl.styles import Font


def stringToRowName(stringData):
    if stringData == 'Users':
        return "C"
    elif stringData == 'Count':
        return "D"
    else:
        return "E"


def fill_value(j, k, current_excel, sheet):
    sheet['A' + str(k)] = current_excel['A' + str(j)].value
    sheet['B' + str(k)] = current_excel['B' + str(j)].value
    sheet[stringToRowName(current_excel['C1'].value) + str(k)] = current_excel['C' + str(j)].value

    return sheet


def withdraw_merge():
    workbook = openpyxl.load_workbook('summary.xlsx')
    sheet = workbook["by currency withdraw"]
    path = 'withdraw'
    data_excel_list = os.listdir(path)

    for i in data_excel_list:
        currentExcel = openpyxl.load_workbook(path + '/' + i).worksheets[0]
        for j in range(2, currentExcel.max_row + 1):
            for k in range(2, sheet.max_row + 2):
                if sheet['A' + str(k)].value is None:
                    sheet = fill_value(j, k, currentExcel, sheet)
                    break
                if sheet['A' + str(k)].value == currentExcel['A' + str(j)].value and \
                        sheet['B' + str(k)].value == currentExcel['B' + str(j)].value:
                    sheet[stringToRowName(currentExcel['C1'].value) + str(k)] = currentExcel['C' + str(j)].value
                    break
                if int(sheet['A' + str(k)].value) > int(currentExcel['A' + str(j)].value):
                    sheet.insert_rows(k)
                    sheet = fill_value(j, k, currentExcel, sheet)
                    break

    font = Font(u'Tahoma', size=8)

    # 將空白補0，數值格式化
    for row in range(1, sheet.max_row + 1):
        if sheet["C" + str(row)].value is None:
            sheet["C" + str(row)] = 0
        if sheet["D" + str(row)].value is None:
            sheet["D" + str(row)] = 0
        if sheet["E" + str(row)].value is None:
            sheet["E" + str(row)] = 0
        sheet["C" + str(row)].font = font
        sheet["D" + str(row)].font = font
        sheet["E" + str(row)].font = font

        if row != 1:
            sheet["C" + str(row)].number_format = '#,##'
            sheet["D" + str(row)].number_format = '#,##'
            sheet["E" + str(row)].number_format = '#,##0.00'

    workbook.save('summary.xlsx')
