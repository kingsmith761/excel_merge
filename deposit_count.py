import numpy as np
import pandas as pd
import openpyxl as op
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font


def deposit_count():
    # df = pd.read_excel('summary.xlsx')
    # pt = pd.pivot_table(df, index='DayOfMonth', columns='Currency', aggfunc={'Count': np.sum})
    #
    # with pd.ExcelWriter('summary.xlsx', mode='a') as writer:
    #     pt.to_excel(writer, sheet_name="d_count")

    load_excel = op.load_workbook('summary.xlsx')
    data_sheet = load_excel["d_cnt.amnt"]
    if "d_count" in load_excel.sheetnames:
        sheet = load_excel["d_count"]
    else:
        sheet = load_excel.create_sheet("d_count")
        sheet["B1"] = "IDR"
        sheet["C1"] = "THB"
        sheet["D1"] = "VND"
        # sheet["E1"] = "Grand Total"
    # sheet.delete_rows(3)

    max_day = data_sheet["A" + str(data_sheet.max_row - 1)].value

    for i in range(2, data_sheet.max_row + 1):
        if int(data_sheet["A" + str(i)].value) <= int(max_day):
            col = int(data_sheet["A" + str(i)].value)
            sheet["A" + str(col + 1)] = str(col)
            if data_sheet["B" + str(i)].value == "IDR":
                sheet["B" + str(col + 1)] = data_sheet["C" + str(i)].value
            elif data_sheet["B" + str(i)].value == "THB":
                sheet["C" + str(col + 1)] = data_sheet["C" + str(i)].value
            elif data_sheet["B" + str(i)].value == "VND":
                sheet["D" + str(col + 1)] = data_sheet["C" + str(i)].value

    chart = LineChart()
    chart.title = "Transaction count"
    # chart.style = 14
    font_test = Font(typeface='Cambria')
    cp = CharacterProperties(latin=font_test, sz=1400)
    chart.y_axis.title = "DayOfMonth"
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.x_axis.title = "Sum of Count"
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

    data = Reference(sheet, min_col=2, min_row=1, max_col=sheet.max_column, max_row=sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    s_idr = chart.series[0]
    s_idr.marker.symbol = "circle"
    s_idr.marker.graphicalProperties.solidFill = "BE4B48"
    s_idr.marker.graphicalProperties.line.solidFill = "BE4B48"

    sheet.add_chart(chart, "N3")
    load_excel.save("summary.xlsx")

# # 設定第 1 條線的樣式
# s1 = chart.series[0]
# s1.marker.symbol = "triangle" # 三角形
# s1.marker.graphicalProperties.solidFill = "FF0000" # 填滿顏色
# s1.marker.graphicalProperties.line.solidFill = "FF0000" # 外框線條顏色
# s1.graphicalProperties.line.noFill = True
#
# # 設定第 2 條線的樣式
# s2 = chart.series[1]
# s2.graphicalProperties.line.solidFill = "00AAAA"
# s2.graphicalProperties.line.dashStyle = "sysDot"
# s2.graphicalProperties.line.width = 100050 # 線條寬度，單位為 EMUs
#
# # 設定第 3 條線的樣式
# s3 = chart.series[2]
# s3.smooth = True # 讓線條平滑
