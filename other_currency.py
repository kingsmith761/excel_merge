import openpyxl as op
from openpyxl.styles import Font


def other_currency():
    load_excel = op.load_workbook('summary.xlsx')
    data_sheet = load_excel["by currency"]
    font = Font(u'Tahoma', size=8)
    if "other market" in load_excel.sheetnames:
        sheet = load_excel["other market"]
    else:
        sheet = load_excel.create_sheet("other market")
        sheet.column_dimensions['A'].width = 20.0
        sheet.column_dimensions['B'].width = 20.0
        sheet.column_dimensions['C'].width = 20.0
        sheet.column_dimensions['E'].width = 20.0
        sheet.column_dimensions['F'].width = 20.0
        sheet.column_dimensions['G'].width = 20.0
        sheet.column_dimensions['I'].width = 20.0
        sheet.column_dimensions['J'].width = 20.0
        sheet.column_dimensions['K'].width = 20.0
        sheet.column_dimensions['M'].width = 20.0
        sheet.column_dimensions['N'].width = 20.0
        sheet.column_dimensions['O'].width = 20.0
        sheet['A1'] = 'MYR'
        sheet['A2'] = 'DayOfMonth'
        sheet['B2'] = 'Count'
        sheet['C2'] = 'Volume'
        sheet['E1'] = 'KRW'
        sheet['E2'] = 'DayOfMonth'
        sheet['F2'] = 'Count'
        sheet['G2'] = 'Volume'
        sheet['I1'] = 'INR'
        sheet['I2'] = 'DayOfMonth'
        sheet['J2'] = 'Count'
        sheet['K2'] = 'Volume'
        sheet['M1'] = 'MMK'
        sheet['M2'] = 'DayOfMonth'
        sheet['N2'] = 'Count'
        sheet['O2'] = 'Volume'

    for i in range(2, data_sheet.max_row + 1):
        pos = data_sheet['A' + str(i)].value
        sheet['A' + str(pos + 2)] = pos
        sheet['E' + str(pos + 2)] = pos
        sheet['I' + str(pos + 2)] = pos
        sheet['M' + str(pos + 2)] = pos
        if data_sheet['B' + str(i)].value == 'MYR':
            sheet['B' + str(pos + 2)] = data_sheet['D' + str(i)].value
            sheet['C' + str(pos + 2)] = data_sheet['E' + str(i)].value
        elif data_sheet['B' + str(i)].value == 'KRW':
            sheet['F' + str(pos + 2)] = data_sheet['D' + str(i)].value
            sheet['G' + str(pos + 2)] = data_sheet['E' + str(i)].value
        elif data_sheet['B' + str(i)].value == 'INR':
            sheet['J' + str(pos + 2)] = data_sheet['D' + str(i)].value
            sheet['K' + str(pos + 2)] = data_sheet['E' + str(i)].value
        elif data_sheet['B' + str(i)].value == 'MMK':
            sheet['N' + str(pos + 2)] = data_sheet['D' + str(i)].value
            sheet['O' + str(pos + 2)] = data_sheet['E' + str(i)].value

    max_row = pos + 2
    sheet['A' + str(max_row + 1)] = 'Grand Total'
    sheet['B' + str(max_row + 1)] = '=SUM(B3:B' + str(max_row) + ')'
    sheet['C' + str(max_row + 1)] = '=SUM(C3:C' + str(max_row) + ')'
    sheet['E' + str(max_row + 1)] = 'Grand Total'
    sheet['F' + str(max_row + 1)] = '=SUM(F3:F' + str(max_row) + ')'
    sheet['G' + str(max_row + 1)] = '=SUM(G3:G' + str(max_row) + ')'
    sheet['I' + str(max_row + 1)] = 'Grand Total'
    sheet['J' + str(max_row + 1)] = '=SUM(J3:J' + str(max_row) + ')'
    sheet['K' + str(max_row + 1)] = '=SUM(K3:K' + str(max_row) + ')'
    sheet['M' + str(max_row + 1)] = 'Grand Total'
    sheet['N' + str(max_row + 1)] = '=SUM(N3:N' + str(max_row) + ')'
    sheet['O' + str(max_row + 1)] = '=SUM(O3:O' + str(max_row) + ')'

    for font_format in range(1, sheet.max_row + 1):
        sheet['A' + str(font_format)].font = font
        sheet['B' + str(font_format)].font = font
        sheet['C' + str(font_format)].font = font
        sheet['E' + str(font_format)].font = font
        sheet['F' + str(font_format)].font = font
        sheet['G' + str(font_format)].font = font
        sheet['I' + str(font_format)].font = font
        sheet['J' + str(font_format)].font = font
        sheet['K' + str(font_format)].font = font
        sheet['M' + str(font_format)].font = font
        sheet['N' + str(font_format)].font = font
        sheet['O' + str(font_format)].font = font

    load_excel.save('summary.xlsx')
