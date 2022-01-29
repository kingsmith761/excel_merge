import openpyxl as op
from openpyxl.styles import Font


def main_deposit():
    load_excel = op.load_workbook('summary.xlsx')
    data_sheet = load_excel["by currency"]
    font = Font(u'Tahoma', size=8)
    if "main" in load_excel.sheetnames:
        sheet = load_excel["main"]
    else:
        sheet = load_excel.create_sheet("main")
        sheet.column_dimensions['A'].width = 20.0
        sheet.column_dimensions['B'].width = 20.0
        sheet.column_dimensions['C'].width = 20.0
        sheet.column_dimensions['D'].width = 20.0
        sheet.column_dimensions['E'].width = 20.0
        sheet.column_dimensions['G'].width = 20.0
        sheet.column_dimensions['H'].width = 20.0
        sheet.column_dimensions['I'].width = 20.0
        sheet.column_dimensions['J'].width = 20.0
        sheet.column_dimensions['K'].width = 20.0
        sheet['A1'] = 'Count'
        sheet['A2'] = 'DayOfMonth'
        sheet['B2'] = 'IDR'
        sheet['C2'] = 'THB'
        sheet['D2'] = 'VND'
        sheet['E2'] = 'Grand Total'
        sheet['G1'] = 'Volume'
        sheet['G2'] = 'DayOfMonth'
        sheet['H2'] = 'IDR'
        sheet['I2'] = 'THB'
        sheet['J2'] = 'VND'
        sheet['K2'] = 'Grand Total'

    for i in range(2, data_sheet.max_row + 1):
        pos = data_sheet['A' + str(i)].value
        sheet['A' + str(pos + 2)] = pos
        sheet['G' + str(pos + 2)] = pos
        if data_sheet['B' + str(i)].value == 'IDR':
            sheet['B' + str(pos + 2)] = data_sheet['D' + str(i)].value
            sheet['H' + str(pos + 2)] = data_sheet['E' + str(i)].value
        elif data_sheet['B' + str(i)].value == 'THB':
            sheet['C' + str(pos + 2)] = data_sheet['D' + str(i)].value
            sheet['I' + str(pos + 2)] = data_sheet['E' + str(i)].value
        elif data_sheet['B' + str(i)].value == 'VND':
            sheet['D' + str(pos + 2)] = data_sheet['D' + str(i)].value
            sheet['J' + str(pos + 2)] = data_sheet['E' + str(i)].value

    for j in range(3, pos + 3):
        sheet['E' + str(j)] = '=SUM(B' + str(j) + ':D' + str(j) + ')'
        sheet['K' + str(j)] = '=SUM(H' + str(j) + ':J' + str(j) + ')'

    max_row = pos + 2
    sheet['A' + str(max_row + 1)] = 'Grand Total'
    sheet['B' + str(max_row + 1)] = '=SUM(B3:B' + str(max_row) + ')'
    sheet['C' + str(max_row + 1)] = '=SUM(C3:C' + str(max_row) + ')'
    sheet['D' + str(max_row + 1)] = '=SUM(D3:D' + str(max_row) + ')'
    sheet['E' + str(max_row + 1)] = '=SUM(E3:E' + str(max_row) + ')'
    sheet['G' + str(max_row + 1)] = 'Grand Total'
    sheet['H' + str(max_row + 1)] = '=SUM(H3:H' + str(max_row) + ')'
    sheet['I' + str(max_row + 1)] = '=SUM(I3:I' + str(max_row) + ')'
    sheet['J' + str(max_row + 1)] = '=SUM(J3:J' + str(max_row) + ')'
    sheet['K' + str(max_row + 1)] = '=SUM(K3:K' + str(max_row) + ')'

    for font_format in range(1, sheet.max_row + 1):
        sheet['A' + str(font_format)].font = font
        sheet['B' + str(font_format)].font = font
        sheet['C' + str(font_format)].font = font
        sheet['D' + str(font_format)].font = font
        sheet['E' + str(font_format)].font = font
        sheet['G' + str(font_format)].font = font
        sheet['H' + str(font_format)].font = font
        sheet['I' + str(font_format)].font = font
        sheet['J' + str(font_format)].font = font
        sheet['K' + str(font_format)].font = font

    load_excel.save('summary.xlsx')
