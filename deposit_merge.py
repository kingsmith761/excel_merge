import os
import openpyxl


def stringToRowName(stringData):
    if stringData == 'Count':
        return "C"
    elif stringData == 'Amount':
        return "D"
    else:
        return "E"


def fill_value(j, k, current_excel, sheet):
    sheet['A' + str(k)] = current_excel['A' + str(j)].value
    sheet['B' + str(k)] = current_excel['B' + str(j)].value
    sheet[stringToRowName(current_excel['C1'].value) + str(k)] = current_excel['C' + str(j)].value

    return sheet


def deposit_merge():
    workbook = openpyxl.load_workbook('summary.xlsx')
    sheet = workbook.worksheets[0]
    path = 'deposit'
    data_excel_list = os.listdir(path)

    for i in data_excel_list:
        currentExcel = openpyxl.load_workbook(path + '/' + i).worksheets[0]
        for j in range(2, currentExcel.max_row + 1):
            for k in range(2, sheet.max_row + 2):
                if sheet['A' + str(k)].value is None:
                    sheet = fill_value(j, k, currentExcel, sheet)
                    break
                if sheet['A' + str(k)].value == currentExcel['A' + str(j)].value and \
                        sheet['B' + str(k)].value == currentExcel['B' + str(j)].value:
                    sheet[stringToRowName(currentExcel['C1'].value) + str(k)] = currentExcel['C' + str(j)].value
                    break
                if int(sheet['A' + str(k)].value) > int(currentExcel['A' + str(j)].value):
                    sheet.insert_rows(k)
                    sheet = fill_value(j, k, currentExcel, sheet)
                    break

    for row in range(2, sheet.max_row + 1):
        if sheet["C" + str(row)].value is None:
            sheet["C" + str(row)] = 0
        if sheet["D" + str(row)].value is None:
            sheet["D" + str(row)] = 0
        if sheet["E" + str(row)].value is None:
            sheet["E" + str(row)] = 0

    workbook.save('summary.xlsx')
